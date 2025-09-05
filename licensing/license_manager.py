import os, json, base64, platform, uuid, sys
from datetime import datetime
from typing import Tuple, Dict, Any

from cryptography.hazmat.primitives import serialization, hashes
from cryptography.hazmat.primitives.asymmetric import padding

# 앱 디렉터리 및 기본 경로
USER_HOME = os.path.expanduser("~")
APP_DIR = os.path.join(os.getenv("APPDATA") or USER_HOME, "CommunityCrawler")
LICENSE_PATH = os.path.join(APP_DIR, "license.lic")
EXE_DIR = os.path.dirname(os.path.abspath(
    sys.executable if getattr(sys, "frozen", False) else __file__
))
PORTABLE_LICENSE = os.path.join(EXE_DIR, "license.lic")

# ====== 공개키를 배포 시 교체하세요 ======
PUBLIC_PEM = b"""-----BEGIN PUBLIC KEY-----
MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEArwh9qGLUP3alVE/keAHz
dV53lBIEVpzzuvTpi/EPXqufIXdjfGupZbpF8M7yUGtdsD8WGpW27BKuR4FQQmPO
SNp6lIPwKlTvn46Y3R/nHFE9s0WazUyWIa7mkA0DbMhTihP6x7Lq2Y0dmEZUTJm0
mKEzG+YF6RwOEmctHG05YqyK7xZEzSNNXK2m3hSCptf4romsrty5Hh64vsZ1nR4Z
rNc3zdmMO4MZFWlccDQRpgvDmTj/+IqbQnsfMdPy8FoW8Wm/zPhKQQ22J1LXirnX
5NoWhclvGNy2i4llOP26cNrvK+s5juGKJGhWe698LnrQZLMtzT27px/oqS7n14Ya
zwIDAQAB
-----END PUBLIC KEY-----
"""
# ======================================


def machine_id() -> str:
    try:
        if platform.system() == "Windows":
            import winreg
            k = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Cryptography")
            v, _ = winreg.QueryValueEx(k, "MachineGuid")
            return v
        elif platform.system() == "Darwin":
            import subprocess
            out = subprocess.check_output(["ioreg", "-rd1", "-c", "IOPlatformExpertDevice"]).decode(errors="ignore")
            return out.split('IOPlatformUUID" = "')[1].split('"')[0]
        else:
            return open("/etc/machine-id").read().strip()
    except Exception:
        return str(uuid.getnode())


def _b64u_decode(s: str) -> bytes:
    pad = "=" * (-len(s) % 4)
    return base64.urlsafe_b64decode(s + pad)


def _b64u(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).decode().rstrip("=")


def verify_license_text(lic_text: str) -> Tuple[bool, str, Dict[str, Any] | None]:
    try:
        lic = json.loads(lic_text)
        payload_json = _b64u_decode(lic["payload"])
        payload = json.loads(payload_json)
        sig = _b64u_decode(lic["sig"])

        pub = serialization.load_pem_public_key(PUBLIC_PEM)
        pub.verify(sig, payload_json, padding.PKCS1v15(), hashes.SHA256())

        dev = payload.get("dev") or ""
        exp = payload.get("exp") or ""
        if dev and dev != machine_id():
            return False, "등록된 PC가 아닙니다.", None
        if exp:
            if datetime.now().date() > datetime.strptime(exp, "%Y-%m-%d").date():
                return False, "라이선스가 만료되었습니다.", None
        return True, "", payload
    except Exception as e:
        return False, f"라이선스 검증 실패: {e}", None


def sign_license_with_private_pem(priv_pem_path: str, user: str, device_id: str, exp_yyyy_mm_dd: str) -> str:
    payload = {"user": user, "dev": device_id, "exp": exp_yyyy_mm_dd, "ver": 1}
    msg = json.dumps(payload, separators=(",", ":")).encode()
    priv = serialization.load_pem_private_key(open(priv_pem_path, "rb").read(), password=None)
    sig = priv.sign(msg, padding.PKCS1v15(), hashes.SHA256())
    lic = {"payload": _b64u(msg), "sig": _b64u(sig)}
    return json.dumps(lic, ensure_ascii=False)


def load_license_from_disk() -> str | None:
    for path in (LICENSE_PATH, PORTABLE_LICENSE):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
        except Exception:
            continue
    return None


def save_license_to_disk(text: str):
    os.makedirs(APP_DIR, exist_ok=True)
    with open(LICENSE_PATH, "w", encoding="utf-8") as f:
        f.write(text)


def watermark_excel(path: str, payload: Dict[str, Any] | None):
    if not payload:
        return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.create_sheet("_meta")
        ws.sheet_state = "hidden"
        ws["A1"], ws["B1"] = "user", payload.get("user", "")
        ws["A2"], ws["B2"] = "device", payload.get("dev", "") or machine_id()
        ws["A3"], ws["B3"] = "exp", payload.get("exp", "")
        wb.save(path)
    except Exception as e:
        print("워터마크 실패:", e)
